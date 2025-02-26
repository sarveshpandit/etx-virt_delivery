#!/usr/bin/env python3


import os
import re
import csv
import sys
import argparse
from openpyxl import Workbook
from importlib import reload


parser = argparse.ArgumentParser()
parser.add_argument("-p", "--path", help="Path to RVTools csv files", required=True)
parser.add_argument("-o", "--outfile", help="Output file name", required=False)
parser.add_argument("-v", "--verbose", help="Show more verbose output", action="store_true", default=False, required=False)
args = parser.parse_args()

whitelist = ['vInfo', 'vCPU', 'vMemory', 'vDisk', 'vNetwork', 'vHost']

def get_files_in_directory(dir_path, extension):
  if not os.path.exists(dir_path):
      return None
  files_with_extension = [
        file_name for file_name in os.listdir(dir_path)
        if file_name.endswith(extension)
  ]
  return files_with_extension

if __name__ == '__main__':
    patt = "_RVTools_export_all.*"
    repl = ""
    instance = os.path.basename(re.sub(patt, repl, args.path)).lower()
    if args.outfile:
        if args.outfile[-4:] == "xlsx":
            outfile = args.outfile.lower()
        else:
            outfile = f"{args.outfile.lower()}.xlsx"
    else:
        outfile = input(f"Please give an output file to create [ Default: '{instance}.xlsx' ]: ")
        if not outfile:
            outfile = f"{instance}.xlsx"
        elif outfile[-4:] == "xlsx":
            outfile = outfile.lower()
        else:
            outfile = f"{outfile.lower()}.xlsx"
    files = get_files_in_directory(args.path, ".csv")
    if args.verbose:
        print(f"Writing spreadsheel '{args.path}/{outfile}'")
    wb = Workbook()
    del wb[wb.sheetnames[0]] 
    for allowed in whitelist:
        for fname in files:
            # The columns are listed as required in the notebook, they are not always available in the csv output so we add them.
            addReqCols = {'Environment': False, 'Final OS': False, 'Disk Size TB': False, 'VM': False, 'Cluster': False, 'Disk Classification': False}
            rows = 0
            addenv = False
            if allowed in fname:
                if args.verbose:
                    print(f"Adding tab '{allowed}'")
                try:
                    with open(f"{args.path}/{fname}", newline='', encoding='utf8', errors='ignore') as f_input:
                        ws = wb.create_sheet(title=allowed)
                        for row in csv.reader(f_input, delimiter=','):
                            if rows == 0:
                                if ("env" in row) and (allowed == "vInfo" or allowed == "vDisk"):
                                    if args.verbose:
                                        print("Converting the env column to Environment column")
                                    row = ["Environment" if x == "env" else x for x in row]
                                for col in addReqCols.keys():
                                    if col not in row and allowed == "vInfo":
                                        row.append(col)
                                        addReqCols[col] = True
                                        if args.verbose:
                                            print(f"Adding Column Header: {col}")
                            else:
                                for col in addReqCols.keys():
                                    if addReqCols[col] and allowed == "vInfo":
                                        row.append('')
                            ws.append(row)
                            rows += 1
                    if args.verbose:
                        print(f"Successfully wrote tab '{allowed}'")
                except:
                    print(f"Uh oh something went wrong writing tab '{allowed}'")
    try:
        wb.save(f"{args.path}/{outfile}")
        if args.verbose:
            print(f"Successfully wrote '{args.path}/{outfile}'")
    except:
        print(f"Uh oh something went wrong writing '{args.path}/{outfile}'")
    
